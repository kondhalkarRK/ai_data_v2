"""
Generate doc/ASKDB_LLM_vs_BI_Cost_Comparison.xlsx
Leadership cost showcase: GPT small/medium/high vs indicative Power BI / Snowflake.
Run: python doc/_generate_llm_bi_cost_xlsx.py
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule

ROOT = Path(__file__).resolve().parent
OUT = ROOT / "ASKDB_LLM_vs_BI_Cost_Comparison.xlsx"

YELLOW = PatternFill("solid", fgColor="FFF2CC")
HEADER = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE = Font(bold=True, size=14, color="1F4E79")
SECTION = Font(bold=True, size=12, color="1F4E79")
LABEL = Font(size=10)
MONEY = '"$"#,##0.00'
MONEY0 = '"$"#,##0'
PCT = "0%"
INT = "#,##0"
DEC = "#,##0.00"
THIN = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)


def _style_header_row(ws, row: int, cols: int) -> None:
    for c in range(1, cols + 1):
        cell = ws.cell(row, c)
        cell.fill = HEADER
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = THIN


def _input(ws, row: int, label: str, value, fmt=None, note: str = "") -> None:
    ws.cell(row, 1, label).font = LABEL
    cell = ws.cell(row, 2, value)
    cell.fill = YELLOW
    cell.border = THIN
    if fmt:
        cell.number_format = fmt
    if note:
        ws.cell(row, 3, note).font = Font(italic=True, size=9, color="666666")


def build() -> Path:
    wb = Workbook()

    # ── Assumptions ──────────────────────────────────────────────
    ws = wb.active
    ws.title = "Assumptions"
    ws["A1"] = "ASK-DB · LLM vs BI cost comparison — Assumptions"
    ws["A1"].font = TITLE
    ws.merge_cells("A1:D1")
    ws["A2"] = (
        "Yellow cells are editable inputs. All other sheets use formulas. "
        "Capgemini Studio rates are INDICATIVE for showcase — not contractual invoices. "
        "Power BI / Snowflake figures are market placeholders — replace with your quotes."
    )
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A2:D2")
    ws.row_dimensions[2].height = 48

    ws["A4"] = "Volume scenario"
    ws["A4"].font = SECTION
    _input(ws, 5, "Named users", 50, INT, "Org seats")
    _input(ws, 6, "Active users / month", 45, INT, "Who actually ask")
    _input(ws, 7, "Questions / active user / month", 100, INT, "Planning figure")
    ws["A8"] = "Derived monthly questions (active × Q)"
    ws["B8"] = "=B6*B7"
    ws["B8"].number_format = INT
    ws["C8"] = "Secondary scenario (45×100 = 4,500)"
    _input(ws, 9, "Showcase monthly questions", 30000, INT, "Primary slide volume")
    _input(ws, 10, "% questions with narration", 0.20, PCT, "LLM narration path")
    ws["A11"] = "% questions without narration"
    ws["B11"] = "=1-B10"
    ws["B11"].number_format = PCT
    _input(ws, 12, "Months / year", 12, INT)

    ws["A14"] = "Token assumptions (from config/llm_catalog.py)"
    ws["A14"].font = SECTION
    _input(ws, 15, "Tokens / question (no narration)", 2500, INT, "SQL NLQ path")
    _input(ws, 16, "Tokens / question (with narration)", 6000, INT, "SQL + LLM narration")
    ws["A17"] = "Blended tokens / question"
    ws["B17"] = "=B11*B15+B10*B16"
    ws["B17"].number_format = DEC
    ws["C17"] = "Weighted by narration mix"

    ws["A19"] = "GPT tier rates — USD per 1M tokens (indicative Studio)"
    ws["A19"].font = SECTION
    _input(ws, 20, "Small — GPT-5 nano ($/1M)", 0.20, DEC, "openai.gpt-5-nano")
    _input(ws, 21, "Medium — GPT-5 mini ($/1M)", 0.80, DEC, "openai.gpt-5-mini")
    _input(ws, 22, "High — GPT-5.1 ($/1M)", 10.00, DEC, "openai.gpt-5.1")

    ws["A24"] = "Classic BI / warehouse placeholders (replace with quotes)"
    ws["A24"].font = SECTION
    _input(ws, 25, "Power BI Pro $/user/month", 14.00, DEC, "Self-serve viz seat")
    _input(ws, 26, "Power BI Premium capacity $/month", 5000.00, DEC, "Optional shared capacity")
    _input(ws, 27, "Use Premium instead of Pro? (1=yes,0=no)", 0, INT, "0 = Pro×users; 1 = Premium only")
    _input(ws, 28, "Snowflake compute $/month", 2500.00, DEC, "Indicative warehouse credits")
    _input(ws, 29, "Snowflake storage $/month", 200.00, DEC, "Indicative storage")
    _input(ws, 30, "Dashboard build / maint $/month", 1500.00, DEC, "Amortized BI build effort")

    ws["A32"] = "Named ranges (cell map)"
    ws["A32"].font = SECTION
    notes = [
        "B5 NamedUsers  B6 ActiveUsers  B7 QPerUser  B8 DerivedQ  B9 ShowcaseQ",
        "B10 PctNarr  B11 PctNoNarr  B12 Months  B15 TokNoNarr  B16 TokNarr  B17 TokBlend",
        "B20 RateSmall  B21 RateMed  B22 RateHigh",
        "B25–B30 BI/Snowflake placeholders",
        "Note: $1 = 1M tokens ONLY when rate = $1.00/1M. At nano $0.20, $1 buys 5M tokens; at GPT-5.1 $10, $1 buys 100k tokens.",
    ]
    for i, t in enumerate(notes):
        ws.cell(33 + i, 1, t).font = Font(size=9, color="555555")

    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 42
    ws.column_dimensions["D"].width = 20

    # ── Unit economics ───────────────────────────────────────────
    ue = wb.create_sheet("LLM_Unit_Economics")
    ue["A1"] = "LLM unit economics by GPT tier"
    ue["A1"].font = TITLE
    ue["A2"] = "Formulas reference Assumptions. Cost/Q = tokens × ($/1M) / 1,000,000"
    ue["A2"].font = Font(italic=True, size=9, color="666666")

    headers = [
        "Tier",
        "Model",
        "USD / 1M tokens",
        "Tokens per $1",
        "Questions per $1 (no narr)",
        "Questions per $1 (with narr)",
        "Cost / Q (no narr)",
        "Cost / Q (with narr)",
        "Cost / Q (blended)",
    ]
    for c, h in enumerate(headers, 1):
        ue.cell(4, c, h)
    _style_header_row(ue, 4, len(headers))

    tiers = [
        (5, "Small", "GPT-5 nano", "Assumptions!B20"),
        (6, "Medium", "GPT-5 mini", "Assumptions!B21"),
        (7, "High", "GPT-5.1", "Assumptions!B22"),
    ]
    for row, name, model, rate_ref in tiers:
        ue.cell(row, 1, name)
        ue.cell(row, 2, model)
        ue.cell(row, 3, f"={rate_ref}")
        ue.cell(row, 3).number_format = DEC
        # Tokens per $1 = 1e6 / rate
        ue.cell(row, 4, f"=IF(C{row}=0,0,1000000/C{row})")
        ue.cell(row, 4).number_format = INT
        # Q per $1 no narr
        ue.cell(row, 5, f"=IF(OR(C{row}=0,Assumptions!B15=0),0,INT(D{row}/Assumptions!B15))")
        ue.cell(row, 5).number_format = INT
        ue.cell(row, 6, f"=IF(OR(C{row}=0,Assumptions!B16=0),0,INT(D{row}/Assumptions!B16))")
        ue.cell(row, 6).number_format = INT
        # Cost per Q
        ue.cell(row, 7, f"=Assumptions!B15*C{row}/1000000")
        ue.cell(row, 7).number_format = '"$"#,##0.0000'
        ue.cell(row, 8, f"=Assumptions!B16*C{row}/1000000")
        ue.cell(row, 8).number_format = '"$"#,##0.0000'
        ue.cell(row, 9, f"=Assumptions!B17*C{row}/1000000")
        ue.cell(row, 9).number_format = '"$"#,##0.0000'
        for c in range(1, 10):
            ue.cell(row, c).border = THIN

    ue["A9"] = "How to read"
    ue["A9"].font = SECTION
    ue["A10"] = (
        "Questions per $1 (no narr) ≈ how many SQL-only Chat answers $1 buys at that tier. "
        "Blended uses Assumptions narration mix. Prefer Insights/Table/Chart (no LLM narration) to stay near the no-narration column."
    )
    ue["A10"].alignment = Alignment(wrap_text=True)
    ue.merge_cells("A10:I10")
    ue.row_dimensions[10].height = 40

    for i, w in enumerate([10, 14, 14, 14, 16, 16, 14, 14, 14], 1):
        ue.column_dimensions[get_column_letter(i)].width = w

    # ── Monthly volume ───────────────────────────────────────────
    vol = wb.create_sheet("Monthly_Volume_Cost")
    vol["A1"] = "Monthly & annual LLM cost by volume scenario"
    vol["A1"].font = TITLE
    vol["A2"] = (
        "Primary showcase = Assumptions!B9 (default 30,000). "
        "Secondary = derived active×Q (Assumptions!B8)."
    )
    vol["A2"].font = Font(italic=True, size=9, color="666666")

    vol["A4"] = "Scenario A — Showcase volume (primary)"
    vol["A4"].font = SECTION
    vol["A5"] = "Monthly questions"
    vol["B5"] = "=Assumptions!B9"
    vol["B5"].number_format = INT
    vol["B5"].fill = YELLOW

    vh = [
        "Tier",
        "Monthly 100% no-narr",
        "Monthly 100% with-narr",
        "Monthly blended",
        "Annual blended",
        "Cost / answered Q (blended)",
    ]
    for c, h in enumerate(vh, 1):
        vol.cell(7, c, h)
    _style_header_row(vol, 7, len(vh))

    # Cost = Q * tokens * rate / 1e6
    # Rows 8–10 Small/Med/High for scenario A
    for i, (name, rate_cell) in enumerate(
        [("Small", "Assumptions!B20"), ("Medium", "Assumptions!B21"), ("High", "Assumptions!B22")],
        start=8,
    ):
        vol.cell(i, 1, name)
        vol.cell(i, 2, f"=$B$5*Assumptions!$B$15*{rate_cell}/1000000")
        vol.cell(i, 3, f"=$B$5*Assumptions!$B$16*{rate_cell}/1000000")
        vol.cell(i, 4, f"=$B$5*Assumptions!$B$17*{rate_cell}/1000000")
        vol.cell(i, 5, f"=D{i}*Assumptions!$B$12")
        vol.cell(i, 6, f"=IF($B$5=0,0,D{i}/$B$5)")
        for c in range(2, 6):
            vol.cell(i, c).number_format = MONEY
        vol.cell(i, 6).number_format = '"$"#,##0.0000'
        for c in range(1, 7):
            vol.cell(i, c).border = THIN

    vol["A12"] = "Scenario B — Derived volume (45 × 100)"
    vol["A12"].font = SECTION
    vol["A13"] = "Monthly questions"
    vol["B13"] = "=Assumptions!B8"
    vol["B13"].number_format = INT

    for c, h in enumerate(vh, 1):
        vol.cell(15, c, h)
    _style_header_row(vol, 15, len(vh))

    for i, (name, rate_cell) in enumerate(
        [("Small", "Assumptions!B20"), ("Medium", "Assumptions!B21"), ("High", "Assumptions!B22")],
        start=16,
    ):
        vol.cell(i, 1, name)
        vol.cell(i, 2, f"=$B$13*Assumptions!$B$15*{rate_cell}/1000000")
        vol.cell(i, 3, f"=$B$13*Assumptions!$B$16*{rate_cell}/1000000")
        vol.cell(i, 4, f"=$B$13*Assumptions!$B$17*{rate_cell}/1000000")
        vol.cell(i, 5, f"=D{i}*Assumptions!$B$12")
        vol.cell(i, 6, f"=IF($B$13=0,0,D{i}/$B$13)")
        for c in range(2, 6):
            vol.cell(i, c).number_format = MONEY
        vol.cell(i, 6).number_format = '"$"#,##0.0000'
        for c in range(1, 7):
            vol.cell(i, c).border = THIN

    # Sensitivity High tier @ showcase Q
    vol["A20"] = "Sensitivity — High tier @ showcase volume (narration mix)"
    vol["A20"].font = SECTION
    sh = ["Narration %", "Blended tokens/Q", "Monthly LLM cost", "Annual LLM cost"]
    for c, h in enumerate(sh, 1):
        vol.cell(21, c, h)
    _style_header_row(vol, 21, 4)

    mixes = [(22, 0.0), (23, 0.20), (24, 0.50), (25, 1.0)]
    for row, mix in mixes:
        vol.cell(row, 1, mix).number_format = PCT
        vol.cell(row, 1).fill = YELLOW
        # tokens = (1-mix)*2500 + mix*6000
        vol.cell(
            row,
            2,
            f"=(1-A{row})*Assumptions!$B$15+A{row}*Assumptions!$B$16",
        )
        vol.cell(row, 2).number_format = DEC
        vol.cell(
            row,
            3,
            f"=Assumptions!$B$9*B{row}*Assumptions!$B$22/1000000",
        )
        vol.cell(row, 3).number_format = MONEY
        vol.cell(row, 4, f"=C{row}*Assumptions!$B$12")
        vol.cell(row, 4).number_format = MONEY
        for c in range(1, 5):
            vol.cell(row, c).border = THIN

    vol["A27"] = (
        "Talking point: at High tier and 30k Q/month, 0% narration ≈ $750/mo; "
        "100% narration ≈ $1,800/mo. Product default (Insights without LLM narration) stays near the low end."
    )
    vol["A27"].alignment = Alignment(wrap_text=True)
    vol.merge_cells("A27:F27")
    vol.row_dimensions[27].height = 36

    # Chart data for showcase blended
    vol["A29"] = "Chart data — Showcase blended monthly by tier"
    vol["A29"].font = SECTION
    vol["A30"] = "Tier"
    vol["B30"] = "Monthly blended $"
    vol["A31"] = "Small"
    vol["B31"] = "=D8"
    vol["A32"] = "Medium"
    vol["B32"] = "=D9"
    vol["A33"] = "High"
    vol["B33"] = "=D10"
    vol["B31"].number_format = MONEY
    vol["B32"].number_format = MONEY
    vol["B33"].number_format = MONEY

    chart = BarChart()
    chart.type = "col"
    chart.title = "Showcase monthly LLM cost (blended)"
    chart.y_axis.title = "USD / month"
    data = Reference(vol, min_col=2, min_row=30, max_row=33)
    cats = Reference(vol, min_col=1, min_row=31, max_row=33)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    chart.width = 12
    chart.height = 8
    vol.add_chart(chart, "D29")

    for i, w in enumerate([28, 20, 22, 18, 16, 22], 1):
        vol.column_dimensions[get_column_letter(i)].width = w

    # ── BI comparison ────────────────────────────────────────────
    bi = wb.create_sheet("BI_Warehouse_Compare")
    bi["A1"] = "Indicative classic BI / warehouse vs ASK-DB LLM path"
    bi["A1"].font = TITLE
    bi["A2"] = (
        "Not 1:1 feature parity. Classic path = license + warehouse + dashboard effort. "
        "ASK-DB path = existing Postgres + usage-based LLM (SQL-first; narration optional). "
        "Replace yellow BI/Snowflake placeholders with your quotes."
    )
    bi["A2"].alignment = Alignment(wrap_text=True)
    bi.merge_cells("A2:F2")
    bi.row_dimensions[2].height = 48

    bi["A4"] = "Classic BI + warehouse stack (monthly)"
    bi["A4"].font = SECTION
    bh = ["Line item", "Monthly USD", "Notes"]
    for c, h in enumerate(bh, 1):
        bi.cell(5, c, h)
    _style_header_row(bi, 5, 3)

    bi["A6"] = "Power BI Pro (users × $/user)"
    bi["B6"] = "=IF(Assumptions!B27=1,0,Assumptions!B5*Assumptions!B25)"
    bi["C6"] = "If Premium flag=1, Pro line is $0"
    bi["A7"] = "Power BI Premium capacity"
    bi["B7"] = "=IF(Assumptions!B27=1,Assumptions!B26,0)"
    bi["C7"] = "Used when Premium flag=1"
    bi["A8"] = "Snowflake compute"
    bi["B8"] = "=Assumptions!B28"
    bi["C8"] = "Indicative credits"
    bi["A9"] = "Snowflake storage"
    bi["B9"] = "=Assumptions!B29"
    bi["A10"] = "Dashboard build / maintenance (amortized)"
    bi["B10"] = "=Assumptions!B30"
    bi["A11"] = "Classic path TOTAL (monthly)"
    bi["A11"].font = Font(bold=True)
    bi["B11"] = "=SUM(B6:B10)"
    bi["B11"].font = Font(bold=True)
    bi["A12"] = "Classic path TOTAL (annual)"
    bi["B12"] = "=B11*Assumptions!B12"
    for r in range(6, 13):
        bi.cell(r, 2).number_format = MONEY
        for c in range(1, 4):
            bi.cell(r, c).border = THIN

    bi["A14"] = "ASK-DB LLM path (monthly) — showcase volume blended"
    bi["A14"].font = SECTION
    for c, h in enumerate(["Tier", "LLM monthly (blended @ showcase Q)", "Annual", "Notes"], 1):
        bi.cell(15, c, h)
    _style_header_row(bi, 15, 4)

    bi["A16"] = "Small"
    bi["B16"] = "=Monthly_Volume_Cost!D8"
    bi["C16"] = "=Monthly_Volume_Cost!E8"
    bi["D16"] = "On existing Postgres — no Snowflake line assumed"
    bi["A17"] = "Medium"
    bi["B17"] = "=Monthly_Volume_Cost!D9"
    bi["C17"] = "=Monthly_Volume_Cost!E9"
    bi["A18"] = "High"
    bi["B18"] = "=Monthly_Volume_Cost!D10"
    bi["C18"] = "=Monthly_Volume_Cost!E10"
    for r in range(16, 19):
        bi.cell(r, 2).number_format = MONEY
        bi.cell(r, 3).number_format = MONEY
        for c in range(1, 5):
            bi.cell(r, c).border = THIN

    bi["A20"] = "Side-by-side monthly (showcase)"
    bi["A20"].font = SECTION
    for c, h in enumerate(["Path", "Monthly USD", "Annual USD", "Cost / Q (vs showcase Q)"], 1):
        bi.cell(21, c, h)
    _style_header_row(bi, 21, 4)

    bi["A22"] = "Classic BI + Snowflake stack"
    bi["B22"] = "=B11"
    bi["C22"] = "=B12"
    bi["D22"] = "=IF(Assumptions!B9=0,0,B22/Assumptions!B9)"
    bi["A23"] = "ASK-DB LLM — Small blended"
    bi["B23"] = "=B16"
    bi["C23"] = "=C16"
    bi["D23"] = "=IF(Assumptions!B9=0,0,B23/Assumptions!B9)"
    bi["A24"] = "ASK-DB LLM — Medium blended"
    bi["B24"] = "=B17"
    bi["C24"] = "=C17"
    bi["D24"] = "=IF(Assumptions!B9=0,0,B24/Assumptions!B9)"
    bi["A25"] = "ASK-DB LLM — High blended"
    bi["B25"] = "=B18"
    bi["C25"] = "=C18"
    bi["D25"] = "=IF(Assumptions!B9=0,0,B25/Assumptions!B9)"
    for r in range(22, 26):
        bi.cell(r, 2).number_format = MONEY
        bi.cell(r, 3).number_format = MONEY
        bi.cell(r, 4).number_format = '"$"#,##0.0000'
        for c in range(1, 5):
            bi.cell(r, c).border = THIN

    bi["A27"] = "Chart data"
    bi["A28"] = "Path"
    bi["B28"] = "Monthly $"
    bi["A29"] = "Classic BI+DW"
    bi["B29"] = "=B22"
    bi["A30"] = "ASK-DB Small"
    bi["B30"] = "=B23"
    bi["A31"] = "ASK-DB Medium"
    bi["B31"] = "=B24"
    bi["A32"] = "ASK-DB High"
    bi["B32"] = "=B25"
    for r in range(29, 33):
        bi.cell(r, 2).number_format = MONEY

    chart2 = BarChart()
    chart2.type = "col"
    chart2.title = "Monthly cost: Classic BI+DW vs ASK-DB LLM"
    chart2.y_axis.title = "USD / month"
    data2 = Reference(bi, min_col=2, min_row=28, max_row=32)
    cats2 = Reference(bi, min_col=1, min_row=29, max_row=32)
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats2)
    chart2.width = 14
    chart2.height = 9
    bi.add_chart(chart2, "F20")

    bi["A34"] = (
        "Caveat: Classic stack often still needs a data platform you already pay for; "
        "ASK-DB LLM is incremental usage on top of Postgres. "
        "If Snowflake is mandatory for other workloads, add it to both sides or keep it as shared corporate cost."
    )
    bi["A34"].alignment = Alignment(wrap_text=True)
    bi.merge_cells("A34:E34")
    bi.row_dimensions[34].height = 48

    for i, w in enumerate([42, 18, 14, 36], 1):
        bi.column_dimensions[get_column_letter(i)].width = w

    # ── Executive summary ────────────────────────────────────────
    ex = wb.create_sheet("Executive_Summary", 0)
    ex["A1"] = "ASK-DB cost showcase — Executive summary"
    ex["A1"].font = TITLE
    ex["A2"] = (
        "Indicative Capgemini Studio GPT rates · token model from ASK-DB (2,500 no-narr / 6,000 with-narr). "
        "Edit yellow cells on Assumptions — this page updates automatically."
    )
    ex["A2"].font = Font(italic=True, size=9, color="666666")
    ex.merge_cells("A2:F2")

    ex["A4"] = "Key volume inputs"
    ex["A4"].font = SECTION
    ex["A5"] = "Showcase questions / month"
    ex["B5"] = "=Assumptions!B9"
    ex["B5"].number_format = INT
    ex["A6"] = "Derived questions / month (active × Q)"
    ex["B6"] = "=Assumptions!B8"
    ex["B6"].number_format = INT
    ex["A7"] = "Narration mix"
    ex["B7"] = '=TEXT(Assumptions!B10,"0%")&" with narr / "&TEXT(Assumptions!B11,"0%")&" without"'
    ex["A8"] = "Blended tokens / question"
    ex["B8"] = "=Assumptions!B17"
    ex["B8"].number_format = DEC

    ex["A10"] = "LLM monthly cost @ showcase volume (blended)"
    ex["A10"].font = SECTION
    for c, h in enumerate(["Tier", "Model", "$/1M tokens", "Monthly blended", "Annual blended", "Q per $1 (no narr)"], 1):
        ex.cell(11, c, h)
    _style_header_row(ex, 11, 6)

    ex["A12"] = "Small"
    ex["B12"] = "GPT-5 nano"
    ex["C12"] = "=Assumptions!B20"
    ex["D12"] = "=Monthly_Volume_Cost!D8"
    ex["E12"] = "=Monthly_Volume_Cost!E8"
    ex["F12"] = "=LLM_Unit_Economics!E5"
    ex["A13"] = "Medium"
    ex["B13"] = "GPT-5 mini"
    ex["C13"] = "=Assumptions!B21"
    ex["D13"] = "=Monthly_Volume_Cost!D9"
    ex["E13"] = "=Monthly_Volume_Cost!E9"
    ex["F13"] = "=LLM_Unit_Economics!E6"
    ex["A14"] = "High"
    ex["B14"] = "GPT-5.1"
    ex["C14"] = "=Assumptions!B22"
    ex["D14"] = "=Monthly_Volume_Cost!D10"
    ex["E14"] = "=Monthly_Volume_Cost!E10"
    ex["F14"] = "=LLM_Unit_Economics!E7"
    for r in range(12, 15):
        ex.cell(r, 3).number_format = DEC
        ex.cell(r, 4).number_format = MONEY
        ex.cell(r, 5).number_format = MONEY
        ex.cell(r, 6).number_format = INT
        for c in range(1, 7):
            ex.cell(r, c).border = THIN

    ex["A16"] = "Vs classic BI + warehouse (indicative)"
    ex["A16"].font = SECTION
    for c, h in enumerate(["Path", "Monthly", "Annual"], 1):
        ex.cell(17, c, h)
    _style_header_row(ex, 17, 3)
    ex["A18"] = "Classic BI + Snowflake stack"
    ex["B18"] = "=BI_Warehouse_Compare!B22"
    ex["C18"] = "=BI_Warehouse_Compare!C22"
    ex["A19"] = "ASK-DB High blended (showcase)"
    ex["B19"] = "=D14"
    ex["C19"] = "=E14"
    ex["A20"] = "ASK-DB Medium blended (showcase)"
    ex["B20"] = "=D13"
    ex["C20"] = "=E13"
    for r in range(18, 21):
        ex.cell(r, 2).number_format = MONEY
        ex.cell(r, 3).number_format = MONEY
        for c in range(1, 4):
            ex.cell(r, c).border = THIN

    ex["A22"] = "Talking points for leadership"
    ex["A22"].font = SECTION
    tips = [
        "1. SQL-first Chat (no LLM narration) uses ~2,500 tokens/Q — default Insights/Table/Chart path.",
        "2. Narration chip is opt-in; at High tier + 30k Q, 100% narration ≈ $1,800/mo vs ~$750 at 0% narration.",
        "3. $1 buys 1M tokens only at $1/1M pricing; nano ($0.20) → 5M tokens/$1; GPT-5.1 ($10) → 100k tokens/$1.",
        "4. Classic BI cost is mostly fixed seats + warehouse; ASK-DB LLM is usage-based on existing Postgres.",
        "5. Replace yellow BI/Snowflake placeholders with your enterprise quotes before final business case.",
        "6. Rates are Capgemini Studio INDICATIVE — confirm against current gateway pricing.",
    ]
    for i, t in enumerate(tips):
        ex.cell(23 + i, 1, t)
        ex.merge_cells(start_row=23 + i, start_column=1, end_row=23 + i, end_column=6)

    for i, w in enumerate([36, 14, 14, 16, 16, 18], 1):
        ex.column_dimensions[get_column_letter(i)].width = w

    # Freeze panes
    for sheet in wb.worksheets:
        sheet.freeze_panes = "A4"

    wb.save(OUT)
    return OUT


if __name__ == "__main__":
    path = build()
    print(f"Wrote {path}")
